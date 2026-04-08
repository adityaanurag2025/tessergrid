# reporter.py
# Generates a formatted three-sheet Excel report after cleaner.py runs.
# Sheet 1: Summary dashboard  Sheet 2: Row-by-row fix log  Sheet 3: Flagged items.
# Part of ChainFix - Supply Chain Data Cleaning Tool
# Run from project root: python src/reporter.py

import sys
import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cleaner import clean_file

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
C_DARK_BLUE   = "1B3A6B"
C_LIGHT_BLUE  = "D6E4F0"
C_GREEN_TEXT  = "2E7D32"
C_GREEN_BG    = "E8F5E9"
C_ORANGE_TEXT = "E65100"
C_ORANGE_HDR  = "FF6D00"
C_ORANGE_BG   = "FFF3E0"
C_PURPLE_BG   = "F3E5F5"
C_GREY_ROW    = "F8F9FA"
C_WHITE       = "FFFFFF"
C_REMOVED     = "B71C1C"   # dark red for "Removed" status

# ---------------------------------------------------------------------------
# Human-readable labels for each fix_type key
# ---------------------------------------------------------------------------
FIX_LABELS = {
    "duplicate_columns":      "Duplicate Columns",
    "mixed_dates":            "Mixed Date Formats",
    "inconsistent_text":      "Inconsistent Text",
    "empty_rows":             "Empty Rows",
    "erp_footer_rows":        "ERP Footer Rows",
    "duplicate_rows":         "Duplicate Rows",
    "whitespace_cleaned":     "Whitespace Cleaned",
    "numbers_fixed":          "Numbers as Text",
    "total_value_calculated": "Total Value Recalc",
}

REMOVED_TYPES = {"empty_rows", "erp_footer_rows", "duplicate_rows", "duplicate_columns"}


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

# Returns a PatternFill for solid colour cells.
def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


# Returns a Font object. All parameters are optional keyword args.
def font(bold=False, color=None, size=11, name="Calibri"):
    kwargs = dict(bold=bold, size=size, name=name)
    if color:
        kwargs["color"] = color
    return Font(**kwargs)


# Returns a centred Alignment, optionally wrapping text.
def align(wrap=False, horizontal="left", vertical="center"):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


# Returns a thin border on all four sides of a cell.
def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


# Applies fill, font, alignment, and optional border to a single cell.
def style_cell(cell, bg=None, fg=None, bold=False, size=11,
               h_align="left", wrap=False, border=False):
    if bg:
        cell.fill = fill(bg)
    if fg or bold or size != 11:
        cell.font = font(bold=bold, color=fg, size=size)
    cell.alignment = align(wrap=wrap, horizontal=h_align)
    if border:
        cell.border = thin_border()


# Writes a value into ws[row][col] (1-based) and returns the cell.
def write(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    return cell


# Sets the width of every column in ws based on the longest value in each column.
# Adds a small padding so content is never clipped.
def autofit_columns(ws, min_width=10, max_width=55):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except (TypeError, AttributeError):
                pass  # non-stringifiable cell value — skip width calculation
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, min_width), max_width)


# ---------------------------------------------------------------------------
# Sheet 1: Summary
# ---------------------------------------------------------------------------

# Builds the Summary sheet with a title block, metadata, fix counts table,
# and a flagged-items section. Accepts the full report_data dict from clean_file().
def build_summary_sheet(wb, report_data):
    ws = wb.create_sheet("Summary")
    fix_log  = report_data["fix_log"]
    # Exclude Notes from flag counts — optional field, not a data quality issue
    flags = [f for f in report_data["flags"] if f["column"].strip().lower() != "notes"]

    # Count fixes by type
    by_type = defaultdict(int)
    for entry in fix_log:
        by_type[entry["fix_type"]] += 1
    by_type["merged_cells"] = report_data["unmerged_count"]

    rows_removed = report_data["rows_before"] - report_data["rows_after"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    row = 1

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = write(ws, row, 1, "CHAINFIX - FIX SUMMARY REPORT")
    style_cell(c, bg=C_DARK_BLUE, fg=C_WHITE, bold=True, size=14, h_align="center")
    ws.row_dimensions[row].height = 28
    row += 1

    # ── Metadata ───────────────────────────────────────────────────────────
    meta = [
        ("Original File",       os.path.basename(report_data["input_file"])),
        ("Clean File",          os.path.basename(report_data["output_file"])),
        ("Report Generated",    now),
        ("Total Rows (Before)", report_data["rows_before"]),
        ("Total Rows (After)",  report_data["rows_after"]),
        ("Rows Removed",        rows_removed),
    ]
    row += 1  # blank spacer
    for label, value in meta:
        c_label = write(ws, row, 1, label + ":")
        style_cell(c_label, bold=True, h_align="right")
        c_val = write(ws, row, 2, value)
        style_cell(c_val)
        row += 1

    row += 1  # blank spacer

    # ── Fixes Applied section header ───────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = write(ws, row, 1, "FIXES APPLIED")
    style_cell(c, bg=C_LIGHT_BLUE, fg=C_DARK_BLUE, bold=True)
    row += 1

    # Column headers
    for col, hdr in enumerate(["Problem Type", "Count", "Status"], start=1):
        c = write(ws, row, col, hdr)
        style_cell(c, bg=C_DARK_BLUE, fg=C_WHITE, bold=True, h_align="center", border=True)
    row += 1

    # Fix rows (alternating background)
    fix_display_order = [
        "duplicate_columns", "mixed_dates", "inconsistent_text",
        "empty_rows", "erp_footer_rows", "duplicate_rows",
        "whitespace_cleaned", "numbers_fixed", "merged_cells",
        "total_value_calculated",
    ]
    for i, fix_type in enumerate(fix_display_order):
        count  = by_type.get(fix_type, 0)
        label  = FIX_LABELS.get(fix_type, fix_type.replace("_", " ").title())
        status = "Removed" if fix_type in REMOVED_TYPES else "Fixed"
        bg = C_WHITE if i % 2 == 0 else C_GREY_ROW
        status_color = C_REMOVED if status == "Removed" else C_GREEN_TEXT

        c1 = write(ws, row, 1, label)
        style_cell(c1, bg=bg, border=True)
        c2 = write(ws, row, 2, count)
        style_cell(c2, bg=bg, h_align="center", border=True)
        c3 = write(ws, row, 3, status)
        style_cell(c3, bg=bg, fg=status_color, bold=True, h_align="center", border=True)
        row += 1

    row += 1  # blank spacer

    # ── Flagged For Review section ─────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = write(ws, row, 1, "FLAGGED FOR REVIEW")
    style_cell(c, bg=C_ORANGE_BG, fg=C_ORANGE_TEXT, bold=True)
    row += 1

    for col, hdr in enumerate(["Issue", "Count", "Action Needed"], start=1):
        c = write(ws, row, col, hdr)
        style_cell(c, bg=C_ORANGE_HDR, fg=C_WHITE, bold=True, h_align="center", border=True)
    row += 1

    # Separate flags by type
    missing_flags = [f for f in flags if f.get("flag_type") not in ("business_rule", "unknown_enum")]
    rule_flags    = [f for f in flags if f.get("flag_type") == "business_rule"]
    enum_flags    = [f for f in flags if f.get("flag_type") == "unknown_enum"]

    missing_groups = defaultdict(int)
    for f in missing_flags:
        missing_groups[f["column"]] += 1

    rule_groups = defaultdict(int)
    for f in rule_flags:
        rule_groups[f.get("issue", f["column"])] += 1

    enum_groups = defaultdict(int)
    for f in enum_flags:
        enum_groups[f.get("issue", f["column"])] += 1

    has_flags = missing_groups or rule_groups or enum_groups

    if has_flags:
        for col_name, count in missing_groups.items():
            c1 = write(ws, row, 1, f"Missing {col_name}")
            style_cell(c1, bg=C_ORANGE_BG, border=True)
            c2 = write(ws, row, 2, count)
            style_cell(c2, bg=C_ORANGE_BG, h_align="center", border=True)
            c3 = write(ws, row, 3, "Manual entry required")
            style_cell(c3, bg=C_ORANGE_BG, fg=C_ORANGE_TEXT, bold=True, border=True)
            row += 1
        for issue, count in rule_groups.items():
            c1 = write(ws, row, 1, issue)
            style_cell(c1, bg=C_ORANGE_BG, border=True)
            c2 = write(ws, row, 2, count)
            style_cell(c2, bg=C_ORANGE_BG, h_align="center", border=True)
            c3 = write(ws, row, 3, "Business rule violation — review required")
            style_cell(c3, bg=C_ORANGE_BG, fg=C_ORANGE_TEXT, bold=True, border=True)
            row += 1
        for issue, count in enum_groups.items():
            c1 = write(ws, row, 1, issue)
            style_cell(c1, bg=C_ORANGE_BG, border=True)
            c2 = write(ws, row, 2, count)
            style_cell(c2, bg=C_ORANGE_BG, h_align="center", border=True)
            c3 = write(ws, row, 3, "Unknown value — manual review required")
            style_cell(c3, bg=C_ORANGE_BG, fg=C_ORANGE_TEXT, bold=True, border=True)
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = write(ws, row, 1, "No flags — all data is complete.")
        style_cell(c, bg=C_GREEN_BG, fg=C_GREEN_TEXT, bold=True)
        row += 1

    row += 1  # blank spacer

    # ── Totals ─────────────────────────────────────────────────────────────
    total_fixes = sum(by_type.values())
    totals = [
        ("TOTAL FIXES APPLIED",    total_fixes),
        ("TOTAL FLAGS FOR REVIEW", len(flags)),
    ]
    for label, value in totals:
        c1 = write(ws, row, 1, label)
        style_cell(c1, bg=C_DARK_BLUE, fg=C_WHITE, bold=True)
        c2 = write(ws, row, 2, value)
        style_cell(c2, bg=C_DARK_BLUE, fg=C_WHITE, bold=True, h_align="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1

    autofit_columns(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 22


# ---------------------------------------------------------------------------
# Sheet 2: Fix Details
# ---------------------------------------------------------------------------

# Builds the Fix Details sheet with one row per fix applied.
# Rule-Based fixes get a light green background; Claude AI fixes get light purple.
def build_fix_details_sheet(wb, fix_log, unmerged_count):
    ws = wb.create_sheet("Fix Details")

    headers = [
        "Row Number", "Column Name", "Problem Type",
        "Original Value", "Fixed Value", "Action Taken", "Fixed By",
    ]

    # Header row
    for col, hdr in enumerate(headers, start=1):
        c = write(ws, 1, col, hdr)
        style_cell(c, bg=C_DARK_BLUE, fg=C_WHITE, bold=True, h_align="center", border=True)

    data_row = 2

    # Write each fix log entry
    for entry in fix_log:
        fix_type = entry.get("fix_type", "")
        fixed_by = entry.get("fixed_by", "Rule-Based")
        bg = C_PURPLE_BG if fixed_by == "Claude AI" else C_GREEN_BG

        row_num  = entry.get("row", "—")
        col_name = entry.get("column", "—")
        original = entry.get("original", "—")
        fixed    = entry.get("fixed", "—")
        action   = entry.get("action", "—")
        label    = FIX_LABELS.get(fix_type, fix_type.replace("_", " ").title())

        values = [row_num, col_name, label, original, fixed, action, fixed_by]
        for col_idx, val in enumerate(values, start=1):
            c = write(ws, data_row, col_idx, val)
            style_cell(c, bg=bg, border=True, wrap=(col_idx in (4, 5, 6)))
        data_row += 1

    # Add merged cell entries (from openpyxl step — no per-cell log, one row per range)
    for i in range(unmerged_count):
        bg = C_GREEN_BG
        values = ["—", "—", "Merged Cells", "Merged range", "Unmerged", "Unmerged and values copied", "Rule-Based"]
        for col_idx, val in enumerate(values, start=1):
            c = write(ws, data_row, col_idx, val)
            style_cell(c, bg=bg, border=True)
        data_row += 1

    # Legend below the table
    data_row += 1
    c = write(ws, data_row, 1, "Legend:")
    style_cell(c, bold=True)
    data_row += 1
    c = write(ws, data_row, 1, "  Green rows = Rule-Based fix")
    style_cell(c, bg=C_GREEN_BG)
    data_row += 1
    c = write(ws, data_row, 1, "  Purple rows = Claude AI fix")
    style_cell(c, bg=C_PURPLE_BG)

    autofit_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 3: Flagged Items
# ---------------------------------------------------------------------------

# Builds the Flagged Items sheet listing every missing value that needs
# a human to review and fill in manually.
# Notes column is excluded — it is an optional field in supply chain data.
def build_flagged_items_sheet(wb, flags):
    ws = wb.create_sheet("Flagged Items")

    # Filter out Notes — optional field, not a data quality issue
    flags = [f for f in flags if f["column"].strip().lower() != "notes"]

    headers = [
        "Row Number", "Column Name", "Issue Description",
        "Original Value", "Recommended Action",
    ]

    # Header row
    for col, hdr in enumerate(headers, start=1):
        c = write(ws, 1, col, hdr)
        style_cell(c, bg=C_ORANGE_HDR, fg=C_WHITE, bold=True, h_align="center", border=True)

    if not flags:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        c = write(ws, 2, 1, "No flagged items — all data is complete.")
        style_cell(c, bg=C_GREEN_BG, fg=C_GREEN_TEXT, bold=True, h_align="center")
    else:
        for data_row, flag in enumerate(flags, start=2):
            col_name  = flag["column"]
            flag_type = flag.get("flag_type")
            if flag_type == "business_rule":
                issue_desc = flag.get("issue", "Business rule violation")
                original   = flag.get("original", "")
                action     = "Business rule violation — review required"
            elif flag_type == "unknown_enum":
                issue_desc = flag.get("issue", f"Unknown {col_name} value")
                original   = flag.get("original", "")
                action     = f"Unknown {col_name} value — manual review required"
            else:
                issue_desc = "Missing value"
                original   = "[empty]"
                action     = f"Please enter {col_name.lower()} manually"
            values = [flag["row"], col_name, issue_desc, original, action]
            for col_idx, val in enumerate(values, start=1):
                c = write(ws, data_row, col_idx, val)
                is_action_col = col_idx == 5
                style_cell(
                    c,
                    bg=C_ORANGE_BG,
                    fg=C_ORANGE_TEXT if is_action_col else None,
                    bold=is_action_col,
                    border=True,
                )

    autofit_columns(ws)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

# Orchestrates report generation: creates the workbook, builds all three
# sheets, removes the default empty sheet, and saves to output_path.
def generate_report(report_data, output_path="data/output/chainfix_report.xlsx"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    build_summary_sheet(wb, report_data)
    build_fix_details_sheet(wb, report_data["fix_log"], report_data["unmerged_count"])
    build_flagged_items_sheet(wb, report_data["flags"])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    total_fixes = len(report_data["fix_log"]) + report_data["unmerged_count"]
    total_flags = len(report_data["flags"])

    print()
    print("CHAINFIX REPORT GENERATED")
    print("--------------------------")
    print(f"File:              {output_path}")
    print(f"Sheets:            Summary, Fix Details, Flagged Items")
    print(f"Total Fixes Logged: {total_fixes}")
    print(f"Total Flags:        {total_flags}")
    print()


# Runs cleaner.py to get the fix data, then generates the full Excel report.
if __name__ == "__main__":
    filepath = "data/samples/messy_supply_chain_data.xlsx"
    if len(sys.argv) > 1:
        filepath = sys.argv[1]

    print("Running cleaner to collect fix data...")
    report_data = clean_file(filepath)
    generate_report(report_data)
