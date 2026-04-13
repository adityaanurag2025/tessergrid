# scanner.py
# Scans an Excel/CSV file and detects all data quality problems.
# Returns a structured list of problems with exact row/column locations.
# Part of Tessergrid - AI Data Cleaner
# Run from project root: python src/scanner.py

import sys
import re
import pandas as pd
import openpyxl
from collections import defaultdict


# Maps a zero-based column index to an Excel column letter (0 -> A, 1 -> B, etc.)
def col_index_to_letter(index):
    letter = ""
    while True:
        letter = chr(ord("A") + index % 26) + letter
        index = index // 26 - 1
        if index < 0:
            break
    return letter


# Loads the Excel file into a pandas DataFrame, keeping the raw string values.
# Uses the first row as header and treats all cells as strings for inspection.
def load_dataframe(filepath):
    df = pd.read_excel(filepath, dtype=str, keep_default_na=False)
    return df


# Checks for duplicate column names (case-insensitive, ignoring extra whitespace).
# Returns a list of problem dicts describing each duplicate pair found.
def check_duplicate_columns(df):
    problems = []
    headers = list(df.columns)
    normalized = [re.sub(r"\s+", " ", h.strip().lower()) for h in headers]
    seen = {}
    for idx, norm in enumerate(normalized):
        if norm in seen:
            original_idx = seen[norm]
            col_a = col_index_to_letter(original_idx)
            col_b = col_index_to_letter(idx)
            problems.append({
                "type": "DUPLICATE COLUMNS",
                "detail": f'"{headers[original_idx]}" (column {col_a}) and "{headers[idx]}" (column {col_b}) are the same column name',
            })
        else:
            seen[norm] = idx
    return problems


# Checks every cell in date-like columns for mixed date formats.
# A column is considered date-like if its header contains "date" (case-insensitive).
# Detects formats: MM/DD/YYYY, DD/MM/YYYY, YYYY-MM-DD, "Month DD YYYY", "Month DD, YYYY".
def check_mixed_date_formats(df):
    DATE_PATTERNS = [
        (r"^\d{2}/\d{2}/\d{4}$",   "MM/DD/YYYY or DD/MM/YYYY (slash format)"),
        (r"^\d{4}-\d{2}-\d{2}$",   "YYYY-MM-DD (ISO format)"),
        (r"^[A-Za-z]+ \d{1,2} \d{4}$",  "Month DD YYYY (long format)"),
        (r"^[A-Za-z]+ \d{1,2}, \d{4}$", "Month DD, YYYY (long format with comma)"),
        (r"^\d{2}-\d{2}-\d{4}$",   "DD-MM-YYYY (dash format)"),
        (r"^\d{1,2}/\d{1,2}/\d{2}$", "MM/DD/YY (2-digit year)"),
        (r"^\d{4}/\d{2}/\d{2}$",   "YYYY/MM/DD (slash ISO)"),
    ]
    problems = []
    for col_idx, col in enumerate(df.columns):
        if "date" not in col.lower():
            continue
        formats_found = set()
        flagged_cells = []
        for row_idx, val in enumerate(df[col]):
            val = str(val).strip()
            if val == "" or val.lower() == "nan":
                continue
            for pattern, label in DATE_PATTERNS:
                if re.match(pattern, val):
                    formats_found.add(label)
                    flagged_cells.append((row_idx, val, label))
                    break
        if len(formats_found) > 1:
            col_letter = col_index_to_letter(col_idx)
            for row_idx, val, label in flagged_cells:
                excel_row = row_idx + 2  # +1 for header row, +1 for 1-based index
                problems.append({
                    "type": "MIXED DATE FORMATS",
                    "detail": f'Row {excel_row}, Column {col_letter} ({col}): "{val}" — {label}',
                })
    return problems


# Checks text columns for inconsistent values that look like the same thing
# written differently (e.g., "US" vs "USA" vs "United States").
# Groups values by their normalized form (lowercase, alphanumeric only)
# and flags any group with more than one distinct original spelling.
# Skips columns where every value is unique and normalization is meaningless.
# SKU, supplier_code, and customer_id are intentionally NOT skipped because
# they contain real inconsistencies that the cleaner fixes (e.g. cust-1002 → CUST-1002).
def check_inconsistent_text(df):
    SKIP_EXACT = {"order_id", "batch_no", "remarks", "notes", "product_name"}
    problems = []
    for col_idx, col in enumerate(df.columns):
        if col.strip().lower() in SKIP_EXACT:
            continue
        col_letter = col_index_to_letter(col_idx)
        value_groups = defaultdict(list)
        for row_idx, val in enumerate(df[col]):
            val_str = str(val).strip()
            if val_str == "" or val_str.lower() == "nan":
                continue
            normalized = re.sub(r"[^a-z0-9]", "", val_str.lower())
            excel_row = row_idx + 2
            value_groups[normalized].append((excel_row, val_str))
        for normalized, occurrences in value_groups.items():
            unique_values = list(dict.fromkeys(v for _, v in occurrences))
            if len(unique_values) > 1:
                locations = ", ".join(f"Row {r}" for r, _ in occurrences)
                problems.append({
                    "type": "INCONSISTENT TEXT VALUES",
                    "detail": (
                        f'Column {col_letter} ({col}): '
                        f'{unique_values} — all appear to be the same value '
                        f'[{locations}]'
                    ),
                })
    return problems


# Checks every cell for missing values (empty string or NaN).
# Reports the exact row number, column letter, and column name.
def check_missing_values(df):
    problems = []
    for col_idx, col in enumerate(df.columns):
        col_letter = col_index_to_letter(col_idx)
        for row_idx, val in enumerate(df[col]):
            val_str = str(val).strip()
            if val_str == "" or val_str.lower() == "nan":
                excel_row = row_idx + 2
                problems.append({
                    "type": "MISSING VALUES",
                    "detail": f"Row {excel_row}, Column {col_letter} ({col}): empty",
                })
    return problems


# Checks string cells for leading/trailing whitespace or multiple consecutive spaces.
# Skips cells that are empty or NaN.
def check_extra_whitespace(df):
    problems = []
    for col_idx, col in enumerate(df.columns):
        col_letter = col_index_to_letter(col_idx)
        for row_idx, val in enumerate(df[col]):
            val_str = str(val)
            if val_str == "" or val_str.lower() == "nan":
                continue
            has_leading = val_str != val_str.lstrip()
            has_trailing = val_str != val_str.rstrip()
            has_double_space = "  " in val_str
            if has_leading or has_trailing or has_double_space:
                excel_row = row_idx + 2
                issues = []
                if has_leading:
                    issues.append("leading spaces")
                if has_trailing:
                    issues.append("trailing spaces")
                if has_double_space:
                    issues.append("double spaces")
                problems.append({
                    "type": "EXTRA WHITESPACE",
                    "detail": f'Row {excel_row}, Column {col_letter} ({col}): "{val_str}" has {", ".join(issues)}',
                })
    return problems


# Detects cells that contain numbers stored as text strings.
# Uses openpyxl to read the raw cell data_type — if a cell is typed as string ("s")
# but its value parses as a valid number, it is flagged.
# Skips date columns (column name contains "date") to avoid false positives.
# Skips 4-digit year values (e.g. "2024") which are not numbers-as-text.
def check_numbers_as_text(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    problems = []
    for col_idx, col_name in enumerate(headers):
        # Skip date columns entirely — numeric-looking strings in date cols are not a data type issue
        if col_name and "date" in str(col_name).lower():
            continue
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[col_idx]
            val = cell.value
            if val is None:
                continue
            val_str = str(val).strip()
            # Skip 4-digit year values — not a numbers-as-text problem
            if re.match(r"^\d{4}$", val_str):
                continue
            # Flag: cell data type is string but value is a valid number
            if cell.data_type == "s":
                try:
                    float(val_str)
                    col_letter = col_index_to_letter(col_idx)
                    problems.append({
                        "type": "NUMBERS STORED AS TEXT",
                        "detail": f'Row {cell.row}, Column {col_letter} ({col_name}): "{val}" is a number stored as text',
                    })
                except (ValueError, TypeError):
                    pass
    return problems


# Finds completely empty rows — rows where every cell is empty or NaN.
# Reports the Excel row number of each empty row found.
def check_empty_rows(df):
    problems = []
    for row_idx, row in df.iterrows():
        values = [str(v).strip() for v in row]
        if all(v == "" or v.lower() == "nan" for v in values):
            excel_row = row_idx + 2
            problems.append({
                "type": "EMPTY ROWS",
                "detail": f"Row {excel_row} is completely empty",
            })
    return problems


# Finds duplicate rows — rows where all business data values are identical.
# Skips completely empty rows and excludes the Notes column from comparison
# because Notes may differ even when the underlying order data is duplicated.
# Reports which rows are duplicates and the Order ID they share.
def check_duplicate_rows(df):
    problems = []
    # Exclude Notes column (last column) and formula/metadata columns from comparison
    cols_to_check = [c for c in df.columns if c.strip().lower() not in ("notes",)]
    non_empty = df[df.apply(lambda r: any(str(v).strip() not in ("", "nan") for v in r), axis=1)]
    subset = non_empty[cols_to_check]
    duplicated_mask = subset.duplicated(keep=False)
    if duplicated_mask.any():
        dupes = non_empty[duplicated_mask]
        seen_keys = {}
        for idx, row in dupes.iterrows():
            key = tuple(str(row[c]).strip() for c in cols_to_check)
            excel_row = idx + 2
            if key not in seen_keys:
                seen_keys[key] = [excel_row]
            else:
                seen_keys[key].append(excel_row)
        for key, rows in seen_keys.items():
            if len(rows) > 1:
                row_list = " and ".join(f"Row {r}" for r in rows)
                order_id = key[0] if key[0] not in ("", "nan") else "unknown"
                problems.append({
                    "type": "DUPLICATE ROWS",
                    "detail": f"{row_list} are identical (Order ID: {order_id})",
                })
    return problems


# Checks the Excel file for merged cells using openpyxl.
# Merged cells are common in ERP exports and cause silent data loss in pandas.
def check_merged_cells(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    problems = []
    for merged_range in ws.merged_cells.ranges:
        problems.append({
            "type": "MERGED CELLS",
            "detail": f"Cells {merged_range} are merged — will cause issues when processing",
        })
    return problems


# Groups a flat list of problem dicts by their "type" key.
# Returns a defaultdict of {type: [detail_string, ...]}
def group_problems_by_type(all_problems):
    grouped = defaultdict(list)
    for p in all_problems:
        grouped[p["type"]].append(p["detail"])
    return grouped


# Prints the final scan report to stdout in a clean, readable format.
def print_report(filepath, df, grouped, total_problems):
    total_rows = len(df)
    total_cols = len(df.columns)

    print()
    print("=====================================")
    print("    CHAINFIX - DATA SCAN REPORT      ")
    print("=====================================")
    print()
    print(f"FILE: {filepath}")
    print(f"TOTAL ROWS: {total_rows}")
    print(f"TOTAL COLUMNS: {total_cols}")
    print()
    print("PROBLEMS FOUND:")
    print("---------------")

    problem_number = 1
    for problem_type, details in grouped.items():
        print(f"\n{problem_number}. {problem_type}")
        for detail in details:
            print(f"   - {detail}")
        problem_number += 1

    print()
    print("=====================================")
    print(f"  TOTAL PROBLEMS FOUND: {total_problems}")
    print("=====================================")
    print()


# Main entry point — runs all checks and prints the full scan report.
def scan_file(filepath):
    df = load_dataframe(filepath)

    all_problems = []
    all_problems += check_duplicate_columns(df)
    all_problems += check_mixed_date_formats(df)
    all_problems += check_inconsistent_text(df)
    all_problems += check_missing_values(df)
    all_problems += check_extra_whitespace(df)
    all_problems += check_numbers_as_text(filepath)
    all_problems += check_empty_rows(df)
    all_problems += check_duplicate_rows(df)
    all_problems += check_merged_cells(filepath)

    grouped = group_problems_by_type(all_problems)
    total_problems = len(all_problems)
    print_report(filepath, df, grouped, total_problems)


if __name__ == "__main__":
    filepath = "data/samples/messy_supply_chain_data.xlsx"
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    scan_file(filepath)
